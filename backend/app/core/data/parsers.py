"""文件解析器 — 将 Excel / CSV / SQLite 转换为结构化数据。"""
from __future__ import annotations
import csv
import io
import re
import sqlite3
import tempfile
from pathlib import Path
import datetime
from dateutil.relativedelta import relativedelta

import openpyxl

from dataclasses import dataclass, field
from datetime import datetime, date
from typing import Optional

# --------------- Type inference ---------------

_DATE_PATTERNS = [
    re.compile(r"^\d{4}-\d{1,2}-\d{1,2}$"),
    re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$"),
    re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$"),
]

_TIMESTAMP_PATTERNS = [
    re.compile(r"^\d{4}-\d{1,2}-\d{1,2}[T ]\d{1,2}:\d{2}"),
    re.compile(r"^\d{4}/\d{1,2}/\d{1,2}[T ]\d{1,2}:\d{2}"),
]

_BOOL_VALUES = {"true", "false", "yes", "no", "1", "0", "t", "f", "y", "n"}

def _try_int(v) -> bool:
    if isinstance(v, (int,)):
        return True
    if isinstance(v, float):
        return v == int(v)
    try:
        int(str(v).strip())
        return True
    except (ValueError, TypeError):
        return False

def _try_float(v) -> bool:
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).strip())
        return True
    except (ValueError, TypeError):
        return False

def _try_bool(v) -> bool:
    return str(v).strip().lower() in _BOOL_VALUES

def _try_date(v) -> bool:
    if isinstance(v, date) and not isinstance(v, datetime):
        return True
    s = str(v).strip()
    return any(p.match(s) for p in _DATE_PATTERNS)

def _try_timestamp(v) -> bool:
    if isinstance(v, datetime):
        return True
    s = str(v).strip()
    return any(p.match(s) for p in _TIMESTAMP_PATTERNS)

def infer_column_type(values: list, sample_size: int = 100) -> str:
    """从数据采样推断列类型。"""
    non_null = [v for v in values[:sample_size] if v is not None and str(v).strip() != ""]
    if not non_null:
        return "varchar"

    if all(_try_bool(v) for v in non_null) and len(non_null) > 0:
        unique_lower = {str(v).strip().lower() for v in non_null}
        if unique_lower <= _BOOL_VALUES and len(unique_lower) <= 4:
            return "boolean"

    if all(_try_int(v) for v in non_null):
        return "integer"

    if all(_try_float(v) for v in non_null):
        return "float"

    if all(_try_timestamp(v) for v in non_null):
        return "timestamp"

    if all(_try_date(v) for v in non_null):
        return "date"

    return "varchar"

PG_TYPE_MAP = {
    "varchar": "TEXT",
    "integer": "BIGINT",
    "float": "DOUBLE PRECISION",
    "boolean": "BOOLEAN",
    "date": "DATE",
    "timestamp": "TIMESTAMP WITH TIME ZONE",
}

def coerce_value(val, col_type: str):
    """
    将原始值转换为 Python 对象，匹配 PG 列类型。

    修复: 空字符串 "" 统一返回 "" (空字符串)，避免 NOT NULL 约束错误。
    对于日期/时间戳列，若单元格为空则返回当前时间作为默认值（避免 NULL）。
    """
    if val is None:
        return None
    if isinstance(val, datetime) and col_type in ("date", "timestamp"):
        if col_type == "date" and isinstance(val, datetime):
            return val.date()
        return val
    s = str(val).strip()
    if s == "":
        # 返回空字符串，避免数据库存储 NULL
        if col_type in ("date", "timestamp"):
            # 日期/时间戳列返回当前时间作为默认值（避免 NULL）
            return datetime.now()
        return ""
    try:
        if col_type == "integer":
            return int(float(s))
        elif col_type == "float":
            return float(s)
        elif col_type == "boolean":
            s_lower = s.lower()
            if s_lower in ("true", "yes", "1", "t", "y"):
                return True
            if s_lower in ("false", "no", "0", "n", "f", "y"):
                return False
            return None
        elif col_type == "date":
            return _parse_date(s)
        elif col_type == "timestamp":
            return _parse_datetime(s)
        else:
            return s
    except (ValueError, TypeError):
        return s


def _parse_date(s: str) -> date | None:

    """尝试多种格式解析日期字符串。"""

    from datetime import date as _date

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m-%d-%Y", "%d-%m-%Y", "%d/%m/%Y"):

        try:

            return datetime.strptime(s, fmt).date()

        except ValueError:

            continue

    return None



def _parse_datetime(s: str) -> datetime | None:

    """尝试多种格式解析时间戳字符串。"""

    for fmt in (

        "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S",

        "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M",

        "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M",

        "%Y-%m-%d %H:%M.%S.%f",

        "%Y-%m-%dT%H:%M:%S.%f",

    ):

        try:

            return datetime.strptime(s, fmt)

        except ValueError:

            continue

    return None



def sanitize_column_name(name: str) -> str:

    """清洗列名，确保 PG 安全（保留中文字符）。"""

    name = str(name).strip()

    name = re.sub(r"[^a-zA-Z0-9_\u4e00-\u9fff]", "_", name)

    name = re.sub(r"_+", "_", name).strip("_")

    if not name:

        return "unnamed"

    if name[0].isdigit():

        name = f"col_{name}"

    return name[:63]



def sanitize_table_name(name: str) -> str:

    """清洗表名，确保 PG 安全（保留中文字符）。"""

    name = str(name).strip().lower()

    name = re.sub(r"[^a-z0-9_\u4e00-\u9fff]", "_", name)

    name = re.sub(r"_+", "_", name).strip("_")

    return name[:50] if name else "unnamed"



# --------------- Data classes ---------------



@dataclass

class ColumnInfo:

    name: str

    type: str = "varchar"

    nullable: bool = True

    comment: str | None = None



@dataclass

class ParsedTable:

    name: str

    columns: list[ColumnInfo]

    data: list[dict]

    row_count: int



@dataclass

class ParseResult:

    tables: list[ParsedTable] = field(default_factory=list)



# --------------- Parsers ---------------



class FileParser:

    """统一文件解析入口。"""



    async def parse(self, file_path: str, file_type: str, original_name: str | None = None) -> ParseResult:

        ext = file_type.lower().lstrip(".")

        if ext in ("xlsx", "xls"):

            return self._parse_excel(file_path)

        elif ext == "csv":

            return self._parse_csv(file_path, original_name=original_name)

        elif ext in ("sqlite", "db"):

            return self._parse_sqlite(file_path)

        else:

            raise ValueError(f"Unsupported file type: {ext}")



    def _parse_excel(self, file_path: str) -> ParseResult:

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        result = ParseResult()



        for sheet_name in wb.sheetnames:

            ws = wb[sheet_name]

            rows_iter = ws.iter_rows(values_only=True)



            header_row = next(rows_iter, None)

            if header_row is None:

                continue



            raw_columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(header_row)]

            columns = [sanitize_column_name(c) for c in raw_columns]

            seen: dict[str, int] = {}

            for i, c in enumerate(columns):

                if c in seen:

                    seen[c] += 1

                    columns[i] = f"{c}_{seen[c]}"

                else:

                    seen[c] = 0



            all_rows: list[dict] = []

            for row_values in rows_iter:

                row_dict = {}

                for j, val in enumerate(row_values):

                    if j < len(columns):

                        row_dict[columns[j]] = val

                all_rows.append(row_dict)



            col_values_map: dict[str, list] = {c: [] for c in columns}

            for row in all_rows[:100]:

                for c in columns:

                    col_values_map[c].append(row.get(c))



            column_infos = []

            for c in columns:

                col_type = infer_column_type(col_values_map[c])

                has_null = any(v is None or str(v).strip() == "" for v in col_values_map[c])

                column_infos.append(ColumnInfo(name=c, type=col_type, nullable=has_null))



            result.tables.append(ParsedTable(

                name=sanitize_table_name(sheet_name),

                columns=column_infos,

                data=all_rows,

                row_count=len(all_rows),

            ))



        wb.close()

        return result



    def _parse_csv(self, file_path: str, original_name: str | None = None) -> ParseResult:

        with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:

            sample = f.read(8192)

            f.seek(0)

            try:

                dialect = csv.Sniffer().sniff(sample)

            except csv.Error:

                dialect = csv.excel

            reader = csv.reader(f, dialect)

            header = next(reader, None)

            if header is None:

                return ParseResult()



            columns = [sanitize_column_name(c) for c in header]

            seen: dict[str, int] = {}

            for i, c in enumerate(columns):

                if c in seen:

                    seen[c] += 1

                    columns[i] = f"{c}_{seen[c]}"

                else:

                    seen[c] = 0



            all_rows: list[dict] = []

            for row_values in reader:

                row_dict = {}

                for j, val in enumerate(row_values):

                    if j < len(columns):

                        val = val.strip() if val else None

                        row_dict[columns[j]] = val

                all_rows.append(row_dict)



            name_source = original_name or file_path

            table_name = sanitize_table_name(Path(name_source).stem)



            col_values_map: dict[str, list] = {c: [] for c in columns}

            for row in all_rows[:100]:

                for c in columns:

                    col_values_map[c].append(row.get(c))



            column_infos = []

            for c in columns:

                col_type = infer_column_type(col_values_map[c])

                has_null = any(v is None or str(v).strip() == "" for v in col_values_map[c])

                column_infos.append(ColumnInfo(name=c, type=col_type, nullable=has_null))



            return ParseResult(tables=[

                ParsedTable(

                    name=table_name,

                    columns=column_infos,

                    data=all_rows,

                    row_count=len(all_rows),

                )

            ])



    def _parse_sqlite(self, file_path: str) -> ParseResult:

        conn = sqlite3.connect(file_path)

        conn.row_factory = sqlite3.Row

        result = ParseResult()



        cursor = conn.execute(

            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"

        )

        table_names = [row[0] for row in cursor.fetchall()]



        for tbl in table_names:

            pragma = conn.execute(f'PRAGMA table_info("{tbl}")').fetchall()

            if not pragma:

                continue



            raw_columns = [row[1] for row in pragma]



            columns = [sanitize_column_name(c) for c in raw_columns]



            rows_cursor = conn.execute(f'SELECT * FROM "{tbl}"')

            raw_rows = rows_cursor.fetchall()



            all_rows: list[dict] = []

            for raw_row in raw_rows:

                row_dict = {}

                for j, c in enumerate(columns):

                    row_dict[c] = raw_row[j] if j < len(raw_row) else None

                all_rows.append(row_dict)



            col_values_map: dict[str, list] = {c: [] for c in columns}

            for row in all_rows[:100]:

                for c in columns:

                    col_values_map[c].append(row.get(c))



            sqlite_types = {row[1].lower(): row[2].upper() for row in pragma}



            column_infos = []

            for i, c in enumerate(columns):

                pg_type = _sqlite_type_to_ours(sqlite_types.get(raw_columns[i].lower(), ""), col_values_map[c])

                notnull = pragma[i][3] == 1

                column_infos.append(ColumnInfo(name=c, type=pg_type, nullable=not notnull))



            result.tables.append(ParsedTable(

                name=sanitize_table_name(tbl),

                columns=column_infos,

                data=all_rows,

                row_count=len(all_rows),

            ))



        conn.close()

        return result



def _sqlite_type_to_ours(sqlite_type: str, values: list) -> str:

    """SQLite 类型 + 数据采样 → 我们的类型。"""

    t = sqlite_type.upper()

    if "INT" in t:

        return "integer"

    if "REAL" in t or "FLOAT" in t or "DOUBLE" in t or "NUMERIC" in t:

        return "float"

    if "BOOL" in t:

        return "boolean"

    if "DATE" in t and "TIME" not in t:

        return "date"

    if "TIME" in t or "DATETIME" in t or "TIMESTAMP" in t:

        return "timestamp"

    return infer_column_type(values)
